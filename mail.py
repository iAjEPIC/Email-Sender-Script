import smtplib
import openpyxl
import time
import random
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.application import MIMEApplication
from email.utils import formatdate
import os
import socket

# Get the directory where the script is located
script_directory = os.path.dirname(os.path.abspath(__file__))

# Define a function to check internet connectivity
def is_connected():
    try:
        socket.create_connection(("8.8.8.8", 53), timeout=5)
        return True
    except OSError:
        pass
    return False

# Function to connect to the SMTP server
def connect_to_server():
    while True:
        if is_connected():
            try:
                server = smtplib.SMTP_SSL(smtp_server, smtp_port)
                server.ehlo()
                server.login(smtp_username, smtp_password)
                print("Connected to SMTP server successfully.")
                return server
            except smtplib.SMTPException as e:
                print(f'Error connecting to SMTP server: {e}')
                server = None
        else:
            print("No internet connection. Retrying in 60 seconds...")
            time.sleep(60)

# Function to send progress notification email
def send_progress_notification(email_count):
    try:
        if is_connected():
            server = connect_to_server()
            if server is not None:
                recipient = "youemail@gmail.com" #Enter you email address here, when 100 mails complete you will received a notification email
                subject = f"{email_count} Emails Sent So Far"
                message = f"{email_count} emails have been sent from your list."
                
                msg = MIMEMultipart()
                msg['From'] = f'{from_name} <{smtp_username}>'
                msg['To'] = recipient
                msg['Subject'] = subject
                msg['Date'] = formatdate(localtime=True)
                
                email_body = MIMEText(message, 'plain')
                msg.attach(email_body)
                
                server.sendmail(smtp_username, [recipient], msg.as_string())
                print(f'Progress notification email sent to {recipient}')
    except smtplib.SMTPException as e:
        print(f'Error sending progress notification email: {e}')

# Load SMTP configuration from Excel
smtp_config_file_path = os.path.join(script_directory, 'smtp.xlsx')

try:
    wb_smtp = openpyxl.load_workbook(smtp_config_file_path)
    ws_smtp = wb_smtp.active
    smtp_username = ws_smtp.cell(row=2, column=2).value
    smtp_password = ws_smtp.cell(row=2, column=3).value
    smtp_server = ws_smtp.cell(row=2, column=4).value
    smtp_port = 465  # Assuming SMTP port is 465 for SSL
    from_name = ws_smtp.cell(row=2, column=1).value  # "From" name from Excel
except Exception as e:
    print(f'Error loading SMTP configuration from Excel: {e}')
    exit()

# Initialize SMTP server
server = None

# Load email addresses from Excel
excel_file_path = os.path.join(script_directory, 'email_list.xlsx')
start_row = 2

try:
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active
    email_list = [ws.cell(row=i, column=1).value for i in range(start_row, ws.max_row + 1)]
except Exception as e:
    print(f'Error loading email addresses from Excel: {e}')
    exit()

# Load blacklist emails
blacklist_emails = ['me.abudhabi@sgs.comTel']

# Read email body HTML content from file
email_body_path = os.path.join(script_directory, 'email.html')

try:
    with open(email_body_path, 'r', encoding='utf-8') as email_body_file:
        email_body_html = email_body_file.read()
except Exception as e:
    print(f'Error reading email body HTML from file: {e}')
    exit()

# Initialize variables
current_row = start_row
emails_sent = 0  # Initialize the counter
progress_notification_interval = 100  # Set the interval for progress notifications

# Load the workbook to update the "Sent" status
wb_email_list = openpyxl.load_workbook(excel_file_path)
ws_email_list = wb_email_list.active

# Define a function to update the delivery status
def update_delivery_status(row, status):
    ws_email_list.cell(row=row, column=2, value=status)
    wb_email_list.save(excel_file_path)

while current_row <= len(email_list):
    server = None  # Reset the server object at the beginning of each iteration
    while not is_connected():
        print("No internet connection. Retrying in 60 seconds...")
        time.sleep(60)
    
    while server is None:
        server = connect_to_server()
        if server is None:
            continue  # Retry the connection until it succeeds

    recipient = email_list[current_row - start_row]

    try:
        if recipient and recipient not in blacklist_emails:
            print(f'{emails_sent + 1} Sending email to {recipient}')  # Print the counter
            emails_sent += 1  # Increment the counter

            msg = MIMEMultipart()
            msg['From'] = f'{from_name} <{smtp_username}>'  # "From" name from Excel
            msg['To'] = recipient
            msg['Subject'] = ws_smtp.cell(row=2, column=6).value  # Subject from Excel
            msg['Date'] = formatdate(localtime=True)  # Add a date header

            # Convert the HTML email body to rich text
            email_body = MIMEText(email_body_html, 'html')

            msg.attach(email_body)

            image_dir = os.path.join(script_directory, 'images')
            image_files = [f for f in os.listdir(image_dir) if f.endswith('.jpg')]

            for i, image_file in enumerate(image_files, start=1):
                with open(os.path.join(image_dir, image_file), "rb") as f:
                    image_data = f.read()
                    msg_image = MIMEImage(image_data)
                    image_cid = f'image{i}'
                    msg_image.add_header('Content-ID', f'<{image_cid}>')
                    msg.attach(msg_image)

            # Add PDF file attachments from the "pdf" directory
            pdf_dir = os.path.join(script_directory, 'pdf')
            pdf_files = [f for f in os.listdir(pdf_dir) if f.endswith('.pdf')]

            for pdf_file in pdf_files:
                pdf_path = os.path.join(pdf_dir, pdf_file)
                with open(pdf_path, "rb") as pdf:
                    pdf_attachment = MIMEApplication(pdf.read(), _subtype="pdf")
                    pdf_attachment.add_header(
                        "Content-Disposition", f"attachment; filename={pdf_file}"
                    )
                    msg.attach(pdf_attachment)

            try:
                server.sendmail(smtp_username, [recipient], msg.as_string())
                print(f'Successfully sent to {recipient}')
                
                # Update "Delivered" status in Excel sheet
                update_delivery_status(current_row, 'Delivered')
                
                # Check if it's time to send a progress notification
                if emails_sent % progress_notification_interval == 0:
                    send_progress_notification(emails_sent)
                
                # Random delay and countdown timer
                random_delay = random.randint(20, 40)
                print(f'Next email will be sent in {random_delay} seconds.')
                for i in range(random_delay, 0, -1):
                    print(f'Sending email in {i} seconds...', end='\r')
                    time.sleep(1)
                print(' ' * 50, end='\r')  # Clear the countdown timer

                current_row += 1
            except smtplib.SMTPException as e:
                print(f'Error sending to {recipient}: {e}')
                # Update "Error" status in Excel sheet
                update_delivery_status(current_row, 'Error')
                current_row += 1  # Continue to the next recipient

        else:
            print(f'Skipping blacklisted email: {recipient}')
            # Update "Not Delivered" status in Excel sheet
            update_delivery_status(current_row, 'Not Delivered')
            current_row += 1

    except Exception as e:
        print(f'Error processing {recipient}: {e}')
        # Update "Error" status in Excel sheet
        update_delivery_status(current_row, 'Error')
        current_row += 1

# End of script
